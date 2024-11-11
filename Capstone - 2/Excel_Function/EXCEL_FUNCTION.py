from openpyxl import load_workbook

class ExcelFunctions:
    def __init__(self, file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    def read_data(self, row, col):
        """Read data from a specific cell.

        Args:
            row (int): Row number (1-indexed).
            col (int): Column number (1-indexed).

        Returns:
            The value of the specified cell, or None if empty.
        """
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row, column=col).value

    def write_data(self, row, col, data):
        """Write data to a specific cell.

        Args:
            row (int): Row number (1-indexed).
            col (int): Column number (1-indexed).
            data: The data to be written.
        """
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row, column=col).value = data
        workbook.save(self.file)